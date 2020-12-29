# -*- coding: utf-8 -*-
from __future__ import absolute_import, print_function, division


import sys
from datetime import datetime
from tempfile import NamedTemporaryFile

import petl as etl
from petl.io.xlsx import fromxlsx, toxlsx, appendxlsx
from petl.test.helpers import ieq, eq_


def _get_test_xlsx():
    try:
        import pkg_resources
        return pkg_resources.resource_filename('petl', 'test/resources/test.xlsx')
    except:
        return None


try:
    # noinspection PyUnresolvedReferences
    import openpyxl
except ImportError as e:
    print('SKIP xlsx tests: %s' % e, file=sys.stderr)
else:

    def test_fromxlsx():
        filename = _get_test_xlsx()
        if filename is None:
            return
        tbl = fromxlsx(filename, 'Sheet1')
        expect = (('foo', 'bar'),
                  ('A', 1),
                  ('B', 2),
                  ('C', 2),
                  (u'é', datetime(2012, 1, 1)))
        ieq(expect, tbl)
        ieq(expect, tbl)

    def test_fromxlsx_read_only():
        filename = _get_test_xlsx()
        if filename is None:
            return
        tbl = fromxlsx(filename, sheet='Sheet1', read_only=True)
        expect = (('foo', 'bar'),
                  ('A', 1),
                  ('B', 2),
                  ('C', 2),
                  (u'é', datetime(2012, 1, 1)))
        ieq(expect, tbl)
        ieq(expect, tbl)

    def test_fromxlsx_nosheet():
        filename = _get_test_xlsx()
        if filename is None:
            return
        tbl = fromxlsx(filename)
        expect = (('foo', 'bar'),
                  ('A', 1),
                  ('B', 2),
                  ('C', 2),
                  (u'é', datetime(2012, 1, 1)))
        ieq(expect, tbl)
        ieq(expect, tbl)

    def test_fromxlsx_range():
        filename = _get_test_xlsx()
        if filename is None:
            return
        tbl = fromxlsx(filename, 'Sheet2', range_string='B2:C6')
        expect = (('foo', 'bar'),
                  ('A', 1),
                  ('B', 2),
                  ('C', 2),
                  (u'é', datetime(2012, 1, 1)))
        ieq(expect, tbl)
        ieq(expect, tbl)

    def test_fromxlsx_offset():
        filename = _get_test_xlsx()
        if filename is None:
            return
        tbl = fromxlsx(filename, 'Sheet1', min_row=2, min_col=2)
        expect = ((1,),
                  (2,),
                  (2,),
                  (datetime(2012, 1, 1, 0, 0),))
        ieq(expect, tbl)
        ieq(expect, tbl)

    def test_toxlsx_appendxlsx():

        # setup
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        # test toxlsx
        toxlsx(tbl, f.name, 'Sheet1')
        actual = fromxlsx(f.name, 'Sheet1')
        ieq(tbl, actual)

        # test appendxlsx
        appendxlsx(tbl, f.name, 'Sheet1')
        expect = etl.cat(tbl, tbl)
        ieq(expect, actual)

    def test_toxlsx_nosheet():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()
        toxlsx(tbl, f.name)
        actual = fromxlsx(f.name)
        ieq(tbl, actual)

    def test_integration():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()
        tbl = etl.wrap(tbl)
        tbl.toxlsx(f.name, 'Sheet1')
        actual = etl.fromxlsx(f.name, 'Sheet1')
        ieq(tbl, actual)
        tbl.appendxlsx(f.name, 'Sheet1')
        expect = tbl.cat(tbl)
        ieq(expect, actual)

    def test_toxlsx_overwrite():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=False, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(1, len(wb.sheetnames))

    def test_toxlsx_replace_file():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        toxlsx(tbl, f.name, sheet=None, mode="replace")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(1, len(wb.sheetnames))

    def test_toxlsx_replace_sheet():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        toxlsx(tbl, f.name, 'Sheet1', mode="replace")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(1, len(wb.sheetnames))

    def test_toxlsx_replace_sheet_nofile():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="replace")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(1, len(wb.sheetnames))

    def test_toxlsx_add_nosheet():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        toxlsx(tbl, f.name, None, mode="add")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(2, len(wb.sheetnames))

    def test_toxlsx_add_sheet_nomatch():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        toxlsx(tbl, f.name, 'Sheet2', mode="add")
        wb = openpyxl.load_workbook(f.name, read_only=True)
        eq_(2, len(wb.sheetnames))

    def test_toxlsx_add_sheet_match():
        tbl = (('foo', 'bar'),
               ('A', 1),
               ('B', 2),
               ('C', 2),
               (u'é', datetime(2012, 1, 1)))
        f = NamedTemporaryFile(delete=True, suffix='.xlsx')
        f.close()

        toxlsx(tbl, f.name, 'Sheet1', mode="overwrite")
        try:
            toxlsx(tbl, f.name, 'Sheet1', mode="add")
            assert False, 'Adding duplicate sheet name did not fail'
        except ValueError:
            pass
