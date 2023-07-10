# -*- coding: utf-8 -*-
from __future__ import absolute_import, print_function, division

import itertools

from petl.compat import PY3, text_type
from petl.util.base import Table, data
from petl.io.sources import read_source_from_arg, write_source_from_arg


def fromxlsx(filename, sheet=None, range_string=None, min_row=None,
             min_col=None, max_row=None, max_col=None, read_only=False,
             **kwargs):
    """
    Extract a table from a sheet in an Excel .xlsx file.

    N.B., the sheet name is case sensitive.

    The `sheet` argument can be omitted, in which case the first sheet in
    the workbook is used by default.

    The `range_string` argument can be used to provide a range string
    specifying a range of cells to extract.

    The `min_row`, `min_col`, `max_row` and `max_col` arguments can be
    used to limit the range of cells to extract. They will be ignored
    if `range_string` is provided.

    The `read_only` argument determines how openpyxl returns the loaded 
    workbook. Default is `False` as it prevents some LibreOffice files
    from getting truncated at 65536 rows. `True` should be faster if the
    file use is read-only and the files are made with Microsoft Excel.

    Any other keyword arguments are passed through to
    :func:`openpyxl.load_workbook()`.

    """

    return XLSXView(filename, sheet=sheet, range_string=range_string,
                    min_row=min_row, min_col=min_col, max_row=max_row,
                    max_col=max_col, read_only=read_only, **kwargs)


class XLSXView(Table):
    
    def __init__(self, filename, sheet=None, range_string=None,
                 min_row=None, min_col=None, max_row=None, max_col=None, 
                 read_only=False, **kwargs):
        self.filename = filename
        self.sheet = sheet
        self.range_string = range_string
        self.min_row = min_row
        self.min_col = min_col
        self.max_row = max_row
        self.max_col = max_col
        self.read_only = read_only
        self.kwargs = kwargs

    def __iter__(self):
        import openpyxl
        source = read_source_from_arg(self.filename)
        with source.open('rb') as source2:
            wb = openpyxl.load_workbook(filename=source2,
                                        read_only=self.read_only,
                                        **self.kwargs)
            if self.sheet is None:
                ws = wb[wb.sheetnames[0]]
            elif isinstance(self.sheet, int):
                ws = wb[wb.sheetnames[self.sheet]]
            else:
                ws = wb[str(self.sheet)]
            if self.range_string is not None:
                rows = ws[self.range_string]
            else:
                rows = ws.iter_rows(min_row=self.min_row,
                                    min_col=self.min_col,
                                    max_row=self.max_row,
                                    max_col=self.max_col)
            for row in rows:
                yield tuple(cell.value for cell in row)
            try:
                wb._archive.close()
            except AttributeError:
                # just here in case openpyxl stops exposing an _archive property.
                pass


def toxlsx(tbl, filename, sheet=None, write_header=True, mode="replace"):
    """
    Write a table to a new Excel .xlsx file.

    N.B., the sheet name is case sensitive.

    The `mode` argument controls how the file and sheet are treated:

      - `replace`: This is the default. It either replaces or adds a
        named sheet, or if no sheet name is provided, all sheets
        (overwrites the entire file).

      - `overwrite`: Always overwrites the file. This produces a file
        with a single sheet.

      - `add`: Adds a new sheet. Raises `ValueError` if a named sheet
        already exists.

    The `sheet` argument can be omitted in all cases. The new sheet
    will then get a default name.
    If the file does not exist, it will be created, unless `replace`
    mode is used with a named sheet. In the latter case, the file
    must exist and be a valid .xlsx file.
    """
    wb = _load_or_create_workbook(filename, mode, sheet)
    ws = _insert_sheet_on_workbook(mode, sheet, wb)
    if write_header:
        it = iter(tbl)
        try:
            hdr = next(it)
            flds = list(map(text_type, hdr))
            rows = itertools.chain([flds], it)
        except StopIteration:
            rows = it
    else:
        rows = data(tbl)
    for row in rows:
        ws.append(row)
    target = write_source_from_arg(filename)
    with target.open('wb') as target2:
        wb.save(target2)


def _load_or_create_workbook(filename, mode, sheet):
    if PY3:
        FileNotFound = FileNotFoundError
    else:
        FileNotFound = IOError

    import openpyxl
    wb = None
    if not (mode == "overwrite" or (mode == "replace" and sheet is None)):
        try:
            source = read_source_from_arg(filename)
            with source.open('rb') as source2:
                wb = openpyxl.load_workbook(filename=source2, read_only=False)
        except FileNotFound:
            wb = None
    if wb is None:
        wb = openpyxl.Workbook(write_only=True)
    return wb


def _insert_sheet_on_workbook(mode, sheet, wb):
    if mode == "replace":
        try:
            ws = wb[str(sheet)]
            ws.delete_rows(1, ws.max_row)
        except KeyError:
            ws = wb.create_sheet(title=sheet)
    elif mode == "add":
        ws = wb.create_sheet(title=sheet)
        # it creates a sheet named "foo1" if "foo" exists.
        if sheet is not None and ws.title != sheet:
            raise ValueError("Sheet %s already exists in file" % sheet)
    elif mode == "overwrite":
        ws = wb.create_sheet(title=sheet)
    else:
        raise ValueError("Unknown mode '%s'" % mode)
    return ws


Table.toxlsx = toxlsx


def appendxlsx(tbl, filename, sheet=None, write_header=False):
    """
    Appends rows to an existing Excel .xlsx file.
    """

    import openpyxl
    source = read_source_from_arg(filename)
    with source.open('rb') as source2:
        wb = openpyxl.load_workbook(filename=source2, read_only=False)
        if sheet is None:
            ws = wb[wb.sheetnames[0]]
        elif isinstance(sheet, int):
            ws = wb[wb.sheetnames[sheet]]
        else:
            ws = wb[str(sheet)]
        if write_header:
            it = iter(tbl)
            try:
                hdr = next(it)
                flds = list(map(text_type, hdr))
                rows = itertools.chain([flds], it)
            except StopIteration:
                rows = it
        else:
            rows = data(tbl)
        for row in rows:
            ws.append(row)
        target = write_source_from_arg(filename)
        with target.open('wb') as target2:
            wb.save(target2)


Table.appendxlsx = appendxlsx
